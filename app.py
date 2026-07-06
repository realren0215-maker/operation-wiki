from __future__ import annotations

import io
import os
import re
import base64
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import TypedDict
from tempfile import NamedTemporaryFile

os.environ.setdefault("MPLCONFIGDIR", "/private/tmp/cpc_matplotlib")

try:
    import fitz
except ImportError:
    fitz = None

import matplotlib
import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image as RLImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

matplotlib.use("Agg")

APP_TITLE = "CPC 광고 운영보고서 자동 생성"
PRIMARY = colors.HexColor("#111111")
ACCENT = colors.HexColor("#FEE500")
ACCENT_SOFT = colors.HexColor("#FFF8CC")
TEXT = colors.HexColor("#1F1F1F")
MUTED = colors.HexColor("#737373")
LIGHT_BG = colors.HexColor("#FAFAF7")
LINE = colors.HexColor("#E8E4D8")
SPEND_COL = "소진비용(위안)"
CPC_COL = "클릭당 평균비용(위안)"
BASE_DIR = Path(__file__).resolve().parent
FONT_DIR = BASE_DIR / "assets" / "fonts"
LOGO_PATH = BASE_DIR / "assets" / "logo.png"
TEMPLATE_DIR = BASE_DIR / "assets" / "templates"
CONTRACT_TEMPLATES = {
    "1년": TEMPLATE_DIR / "contract_1year.xlsx",
    "3개월": TEMPLATE_DIR / "contract_3months.xlsx",
}
CONTRACT_TYPE_LABELS = {
    "1년": "1年",
    "3개월": "3个月",
}
CITY_OPTIONS = ["首尔", "济州岛", "仁川", "釜山", "京畿道"]
YELLOW_FILL_RGB = {"FFFFFF00", "FFFF00", "00FFFF00"}
NOTO_FONT = BASE_DIR / "assets" / "fonts" / "NotoSansKR-Static-Regular.ttf"
NOTO_VAR_FONT = BASE_DIR / "assets" / "fonts" / "NotoSansKR-Regular.ttf"

CANONICAL_COLUMNS = {
    "date": "날짜",
    "impressions": "노출수",
    "clicks": "클릭수",
    "spend": SPEND_COL,
    "cpc": CPC_COL,
    "voucher": "상품권 조회수",
    "save": "저장하기",
    "share": "공유하기",
}

COLUMN_PATTERNS = {
    "date": ["日期", "时间", "日", "date", "day", "날짜", "일자"],
    "impressions": ["曝光（次）", "曝光", "展示", "展现", "impression", "노출"],
    "clicks": ["点击（次）", "点击", "click", "클릭"],
    "spend": ["花费（元）", "花费", "消耗", "消费", "金额", "费用", "cost", "spend", "소진", "비용"],
    "cpc": ["平均点击价格（元）", "平均点击价格", "平均点击", "点击价格", "CPC", "平均价格", "클릭당", "평균비용"],
    "voucher": ["查看团购", "团购", "券", "商品券", "voucher", "product view", "상품권", "조회"],
    "save": ["收藏", "收藏数", "save", "saved", "저장", "저장하기"],
    "share": ["分享", "分享数", "share", "shared", "공유", "공유하기"],
}

REQUIRED_COLUMNS = ["날짜", SPEND_COL, "노출수", "클릭수", CPC_COL, "상품권 조회수", "저장하기", "공유하기"]
APP_NAME = "내부 업무 자동화"


@dataclass(frozen=True)
class NavItem:
    key: str
    label: str
    group: str
    description: str = ""


NAV_ITEMS = [
    NavItem("cpc_report", "CPC 보고서 생성", "업무 자동화", "여러 CPC 파일을 업로드해 PDF/PNG 운영보고서를 생성합니다."),
    NavItem("contract_generator", "계약서 자동 생성", "업무 자동화", "1年/3个月 계약서 템플릿에 입력값을 반영해 Excel/PDF/ZIP 파일을 생성합니다."),
    NavItem("manual_cpc_registration", "CPC 등록 가이드", "업무 매뉴얼"),
    NavItem("manual_permission_transfer", "CPT/CPC 권한 이전 가이드", "업무 매뉴얼"),
    NavItem("manual_backoffice_extract", "백오피스 데이터 추출 가이드", "업무 매뉴얼"),
    NavItem("manual_backoffice_cpc", "백오피스 CPC 등록 방법", "업무 매뉴얼"),
    NavItem("manual_review_delete", "리뷰 삭제 요청 가이드", "업무 매뉴얼"),
    NavItem("manual_faq", "자주 묻는 질문(FAQ)", "업무 매뉴얼"),
    NavItem("manual_search", "검색 기능", "업무 매뉴얼"),
]

NAV_BY_KEY = {item.key: item for item in NAV_ITEMS}
DEFAULT_NAV_KEY = "cpc_report"


@dataclass
class StoreInputs:
    store_name: str
    period_start: date
    period_end: date
    pause_date: date | None
    memo: str
    logo_bytes: bytes | None = None


@dataclass
class ContractInputs:
    contract_type: str
    kadob_no: str
    store_name: str
    address_kr: str
    address_en: str
    business_name_kr: str
    passport_english_name: str
    passport_number: str
    region: str
    store_code: str
    contract_start: date
    cpc_amount: float
    daily_budget: float
    total_amount: float
    memo: str
    signature_bytes: bytes | None = None
    stamp_bytes: bytes | None = None


class ReportFonts(TypedDict):
    medium: str
    bold: str
    extra_bold: str
    chart_path: str | None


def setup_page() -> None:
    st.set_page_config(page_title=APP_NAME, page_icon="📊", layout="wide", initial_sidebar_state="expanded")
    st.markdown(
        f"""
        <style>
        {report_font_face_css()}
        .stApp {{
            background: #f6f7f9;
            font-family: "Pretendard", -apple-system, BlinkMacSystemFont, "Apple SD Gothic Neo", "Noto Sans KR", "Malgun Gothic", sans-serif;
            font-variant-numeric: tabular-nums;
            font-feature-settings: "tnum";
        }}
        [data-testid="stSidebar"] {{
            background: #111827;
        }}
        [data-testid="stSidebar"] .stMarkdown,
        [data-testid="stSidebar"] .stMarkdown p {{
            color: #f9fafb !important;
        }}
        [data-testid="stSidebar"] button p {{
            color: #111827 !important;
        }}
        [data-testid="stSidebar"] button[kind="primary"] p {{
            color: #ffffff !important;
        }}
        .sidebar-title {{
            padding: 10px 4px 18px;
            font-size: 18px;
            font-weight: 800;
            letter-spacing: 0;
        }}
        .sidebar-caption {{
            margin-top: -14px;
            margin-bottom: 18px;
            color: #9ca3af;
            font-size: 13px;
            line-height: 1.5;
        }}
        .app-header {{
            padding: 24px 28px;
            border: 1px solid #e5e7eb;
            border-radius: 12px;
            background: #ffffff;
            margin-bottom: 22px;
        }}
        .app-header .eyebrow {{
            margin: 0 0 8px;
            color: #6b7280;
            font-size: 13px;
            font-weight: 700;
        }}
        .app-header h1 {{
            margin: 0;
            color: #111827;
            font-size: 30px;
            font-weight: 800;
            letter-spacing: 0;
        }}
        .app-header p {{
            margin: 8px 0 0;
            color: #4b5563;
            line-height: 1.55;
        }}
        .coming-soon {{
            padding: 52px 32px;
            border: 1px dashed #cbd5e1;
            border-radius: 12px;
            background: #ffffff;
            text-align: center;
        }}
        .coming-soon h2 {{
            margin: 0;
            color: #111827;
            font-size: 24px;
            font-weight: 800;
            letter-spacing: 0;
        }}
        .coming-soon p {{
            margin: 10px auto 0;
            max-width: 520px;
            color: #6b7280;
            line-height: 1.6;
        }}
        div[data-testid="stMetric"] {{
            background: #ffffff;
            border: 1px solid #e5e7eb;
            border-radius: 10px;
            padding: 12px;
            min-width: 120px;
        }}
        div[data-testid="stMetricValue"] {{ white-space: normal; overflow-wrap: anywhere; font-weight: 800; }}
        .hint {{ color: #4b5563; font-size: 14px; line-height: 1.6; }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_page_header(title: str, description: str, eyebrow: str = APP_NAME) -> None:
    st.markdown(
        f"""
        <div class="app-header">
            <p class="eyebrow">{eyebrow}</p>
            <h1>{title}</h1>
            <p>{description}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_sidebar() -> str:
    st.sidebar.markdown(
        """
        <div class="sidebar-title">내부 업무 자동화</div>
        <div class="sidebar-caption">반복 업무 도구와 매뉴얼을 한 곳에서 관리합니다.</div>
        """,
        unsafe_allow_html=True,
    )
    selected_key = st.session_state.get("active_nav", DEFAULT_NAV_KEY)
    items_by_group: dict[str, list[NavItem]] = {}
    for item in NAV_ITEMS:
        items_by_group.setdefault(item.group, []).append(item)

    for group, items in items_by_group.items():
        st.sidebar.markdown(f"**{group}**")
        for item in items:
            button_type = "primary" if item.key == selected_key else "secondary"
            if st.sidebar.button(item.label, key=f"nav_{item.key}", type=button_type, use_container_width=True):
                selected_key = item.key

    st.session_state["active_nav"] = selected_key
    return selected_key


def normalize_column_name(value: object) -> str:
    text = str(value).strip().lower()
    return re.sub(r"[\s_\-()（）【】\[\]:：/\\]+", "", text)


def score_column(column: str, patterns: list[str]) -> float:
    normalized = normalize_column_name(column)
    best = 0.0
    for pattern in patterns:
        target = normalize_column_name(pattern)
        if target and (target in normalized or normalized in target):
            best = max(best, 1.0)
        best = max(best, SequenceMatcher(None, normalized, target).ratio())
    return best


def build_column_mapping(columns: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    used: set[str] = set()
    for canonical, patterns in COLUMN_PATTERNS.items():
        scored = sorted(((score_column(col, patterns), col) for col in columns if col not in used), reverse=True)
        if not scored:
            continue
        score, source = scored[0]
        threshold = 0.48 if canonical in {"date", "impressions", "clicks", "spend"} else 0.55
        if score >= threshold:
            mapping[source] = CANONICAL_COLUMNS[canonical]
            used.add(source)
    return mapping


def parse_number(value: object) -> float:
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "—", "无", "없음"}:
        return 0.0
    text = text.replace(",", "")
    text = re.sub(r"[元￥₩%\s]", "", text)
    match = re.search(r"-?\d+(\.\d+)?", text)
    return float(match.group()) if match else 0.0


def read_best_sheet(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    xls = pd.ExcelFile(uploaded_file)
    best_sheet = xls.sheet_names[0]
    best_score = -1
    for sheet in xls.sheet_names:
        sample = pd.read_excel(xls, sheet_name=sheet, nrows=20)
        mapping = build_column_mapping([str(col) for col in sample.columns])
        score = len(mapping) + len(sample)
        if score > best_score:
            best_score = score
            best_sheet = sheet
    return pd.read_excel(xls, sheet_name=best_sheet)


def preprocess_dataframe(raw_df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str], list[str]]:
    df = raw_df.copy()
    df.columns = [str(col).strip() for col in df.columns]
    df = df.dropna(how="all")
    mapping = build_column_mapping(list(df.columns))
    df = df.rename(columns=mapping)
    messages: list[str] = []

    if "날짜" not in df.columns:
        possible_date = None
        for col in df.columns:
            parsed = pd.to_datetime(df[col], errors="coerce")
            if parsed.notna().sum() >= max(2, len(df) * 0.4):
                possible_date = col
                break
        if possible_date:
            df = df.rename(columns={possible_date: "날짜"})
            messages.append(f"'{possible_date}' 컬럼을 날짜로 자동 인식했습니다.")
        else:
            df["날짜"] = pd.date_range(datetime.today().date(), periods=len(df), freq="D")
            messages.append("날짜 컬럼을 찾지 못해 행 순서 기준 임시 날짜를 사용했습니다.")

    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce")
    df = df[df["날짜"].notna()].copy()
    for col in REQUIRED_COLUMNS:
        if col not in df.columns:
            df[col] = 0.0
            if col != "날짜":
                messages.append(f"'{col}' 컬럼이 없어 0으로 계산했습니다.")

    for col in [SPEND_COL, "노출수", "클릭수", CPC_COL, "상품권 조회수", "저장하기", "공유하기"]:
        df[col] = df[col].apply(parse_number)

    missing_cpc = (df[CPC_COL] <= 0) & (df["클릭수"] > 0)
    df.loc[missing_cpc, CPC_COL] = df.loc[missing_cpc, SPEND_COL] / df.loc[missing_cpc, "클릭수"]

    agg_rules = {SPEND_COL: "sum", "노출수": "sum", "클릭수": "sum", CPC_COL: "mean", "상품권 조회수": "sum", "저장하기": "sum", "공유하기": "sum"}
    df = df[REQUIRED_COLUMNS].sort_values("날짜").groupby("날짜", as_index=False).agg(agg_rules)
    df[CPC_COL] = df.apply(lambda row: row[SPEND_COL] / row["클릭수"] if row["클릭수"] > 0 else row[CPC_COL], axis=1)
    return df, mapping, messages


def split_latest_months(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame | None, date, date, list[str]]:
    if df.empty:
        today = date.today()
        return df, None, today, today, ["업로드 데이터에서 월별 데이터를 확인하지 못했습니다."]

    working = df.copy()
    working["월"] = working["날짜"].dt.to_period("M")
    months = sorted(working["월"].dropna().unique())
    latest_month = months[-1]
    previous_month = latest_month - 1

    current_df = working[working["월"] == latest_month].drop(columns=["월"]).copy()
    previous_df = working[working["월"] == previous_month].drop(columns=["월"]).copy()
    previous_result = previous_df if not previous_df.empty else None

    period_start = current_df["날짜"].min().date()
    period_end = current_df["날짜"].max().date()
    messages = [f"가장 최근 월({latest_month.strftime('%Y.%m')})을 대상월 데이터로 자동 인식했습니다."]
    if previous_result is not None:
        messages.append(f"직전 월({previous_month.strftime('%Y.%m')})을 전월 비교 데이터로 자동 인식했습니다.")
    else:
        messages.append("직전 월 데이터가 없어 전월 대비 비교 없이 현재월 기준으로 보고서를 생성했습니다.")

    return current_df, previous_result, period_start, period_end, messages


def select_latest_month(df: pd.DataFrame, label: str) -> tuple[pd.DataFrame, date, date, list[str]]:
    if df.empty:
        today = date.today()
        return df, today, today, [f"{label} 데이터에서 월별 데이터를 확인하지 못했습니다."]

    working = df.copy()
    working["월"] = working["날짜"].dt.to_period("M")
    latest_month = sorted(working["월"].dropna().unique())[-1]
    selected = working[working["월"] == latest_month].drop(columns=["월"]).copy()
    period_start = selected["날짜"].min().date()
    period_end = selected["날짜"].max().date()
    return selected, period_start, period_end, [f"{label} 파일의 가장 최근 월({latest_month.strftime('%Y.%m')})을 사용했습니다."]


def latest_month_period(df: pd.DataFrame) -> pd.Period | None:
    if df.empty:
        return None
    months = df["날짜"].dt.to_period("M").dropna().unique()
    if len(months) == 0:
        return None
    return sorted(months)[-1]


def calculate_metrics(df: pd.DataFrame) -> dict[str, float]:
    total_spend = float(df[SPEND_COL].sum()) if not df.empty else 0.0
    total_impressions = float(df["노출수"].sum()) if not df.empty else 0.0
    total_clicks = float(df["클릭수"].sum()) if not df.empty else 0.0
    total_voucher = float(df["상품권 조회수"].sum()) if not df.empty else 0.0
    avg_cpc = total_spend / total_clicks if total_clicks else 0.0
    click_rate = total_clicks / total_impressions if total_impressions else 0.0
    return {
        "total_spend": total_spend,
        "total_impressions": total_impressions,
        "total_clicks": total_clicks,
        "total_voucher": total_voucher,
        "avg_cpc": avg_cpc,
        "click_rate": click_rate,
    }


def format_currency(value: float) -> str:
    return f"{value:,.0f}위안"


def format_cpc_currency(value: float) -> str:
    if abs(value) >= 10:
        return format_currency(value)
    text = f"{value:,.2f}".rstrip("0").rstrip(".")
    return f"{text}위안"


def format_daily_cpc(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "-"
    return f"{value:,.2f}위안"


def format_rate(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "-"
    return f"{value * 100:.2f}%"


def format_number(value: float) -> str:
    return f"{value:,.0f}"


def format_short_date(value: date | datetime | pd.Timestamp | None) -> str:
    if value is None or pd.isna(value):
        return "-"
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        value = value.date()
    return value.strftime("%m.%d")


def format_period(start: date | datetime | pd.Timestamp, end: date | datetime | pd.Timestamp) -> str:
    return f"{format_short_date(start)} ~ {format_short_date(end)}"


def format_month_full(value: date | datetime | pd.Timestamp | None) -> str:
    if value is None or pd.isna(value):
        return "-"
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        value = value.date()
    return f"{value.month}월"


def format_report_title_month(value: date | datetime | pd.Timestamp | None) -> str:
    if value is None or pd.isna(value):
        return "월간"
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        value = value.date()
    return f"{value.month}월"


def comparison_value(current: float, previous: float) -> float | None:
    if previous == 0:
        return 0.0 if current == 0 else None
    return ((current - previous) / previous) * 100


def format_change_percent(value: float | None) -> str:
    if value is None:
        return "-"
    if abs(value) < 0.0001:
        return "0%"
    return f"{'▲' if value > 0 else '▼'} {abs(value):.1f}%"


def build_month_comparison(
    current: dict[str, float],
    previous: dict[str, float] | None,
    current_days: int = 1,
    previous_days: int = 1,
    include_change: bool = True,
) -> list[dict[str, object]]:
    fields = [
        ("노출수", "total_impressions", "count"),
        ("클릭수", "total_clicks", "count"),
        ("광고비", "total_spend", "money"),
        ("클릭당 평균비용(CPC)", "avg_cpc", "cpc"),
    ]
    rows = []
    for label, key, kind in fields:
        current_value = float(current[key])
        previous_value = float(previous[key]) if previous else 0.0
        change = comparison_value(current_value, previous_value) if previous and include_change else None
        if kind == "money":
            current_text = format_currency(current_value)
            previous_text = format_currency(previous_value) if previous else "-"
        elif kind == "cpc":
            current_text = format_cpc_currency(current_value)
            previous_text = format_cpc_currency(previous_value) if previous else "-"
        else:
            current_text = f"{format_number(current_value)}회"
            previous_text = f"{format_number(previous_value)}회" if previous else "-"
        rows.append(
            {
                "label": label,
                "current": current_value,
                "previous": previous_value,
                "current_text": current_text,
                "previous_text": previous_text,
                "change": change,
                "change_text": format_change_percent(change) if include_change else "-",
            }
        )
    return rows


def find_pretendard_font(weight: str) -> Path | None:
    patterns = [
        f"Pretendard-{weight}.ttf",
        f"Pretendard-{weight}.otf",
        f"*Pretendard*{weight}*.ttf",
        f"*Pretendard*{weight}*.otf",
    ]
    for pattern in patterns:
        matches = sorted(FONT_DIR.glob(pattern))
        if matches:
            return matches[0]
    return None


def find_dongle_font(weight: str) -> Path | None:
    patterns = [
        f"Dongle-{weight}.ttf",
        f"Dongle-{weight}.otf",
        f"*Dongle*{weight}*.ttf",
        f"*Dongle*{weight}*.otf",
    ]
    for pattern in patterns:
        matches = sorted(FONT_DIR.glob(pattern))
        if matches:
            return matches[0]
    return None


def font_data_uri(path: Path) -> str:
    mime = "font/ttf" if path.suffix.lower() == ".ttf" else "font/otf"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def report_font_face_css() -> str:
    font_files = {
        "Dongle": {
            "Light": 300,
            "Regular": 400,
            "Bold": 700,
        },
        "Pretendard": {
            "Regular": 400,
            "Medium": 500,
            "SemiBold": 600,
            "Bold": 700,
            "ExtraBold": 800,
        },
    }
    rules = []
    for family, weights in font_files.items():
        for label, weight in weights.items():
            path = find_dongle_font(label) if family == "Dongle" else find_pretendard_font(label)
            if not path:
                continue
            font_format = "truetype" if path.suffix.lower() == ".ttf" else "opentype"
            rules.append(
                f"""
                @font-face {{
                    font-family: '{family}';
                    font-style: normal;
                    font-weight: {weight};
                    src: url('{font_data_uri(path)}') format('{font_format}');
                    font-display: swap;
                }}
                """
            )
    return "\n".join(rules)


def register_pdf_fonts() -> ReportFonts:
    pretendard_regular = find_pretendard_font("Regular")
    pretendard_medium = find_pretendard_font("Medium") or pretendard_regular
    pretendard_semi_bold = find_pretendard_font("SemiBold")
    pretendard_bold = find_pretendard_font("Bold") or pretendard_semi_bold or pretendard_medium
    pretendard_extra_bold = find_pretendard_font("ExtraBold") or pretendard_bold
    if pretendard_medium and pretendard_bold and pretendard_extra_bold:
        pdfmetrics.registerFont(TTFont("Pretendard-Medium", str(pretendard_medium)))
        pdfmetrics.registerFont(TTFont("Pretendard-Bold", str(pretendard_bold)))
        pdfmetrics.registerFont(TTFont("Pretendard-ExtraBold", str(pretendard_extra_bold)))
        pdfmetrics.registerFontFamily(
            "Pretendard",
            normal="Pretendard-Medium",
            bold="Pretendard-Bold",
            italic="Pretendard-Medium",
            boldItalic="Pretendard-Bold",
        )
        return {
            "medium": "Pretendard-Medium",
            "bold": "Pretendard-Bold",
            "extra_bold": "Pretendard-ExtraBold",
            "chart_path": str(pretendard_medium),
        }

    dongle_regular = find_dongle_font("Regular")
    dongle_bold = find_dongle_font("Bold") or dongle_regular
    if dongle_regular and dongle_bold:
        pdfmetrics.registerFont(TTFont("Dongle-Regular", str(dongle_regular)))
        pdfmetrics.registerFont(TTFont("Dongle-Bold", str(dongle_bold)))
        pdfmetrics.registerFontFamily(
            "Dongle",
            normal="Dongle-Regular",
            bold="Dongle-Bold",
            italic="Dongle-Regular",
            boldItalic="Dongle-Bold",
        )
        return {
            "medium": "Dongle-Regular",
            "bold": "Dongle-Bold",
            "extra_bold": "Dongle-Bold",
            "chart_path": str(dongle_regular),
        }

    if NOTO_FONT.exists():
        pdfmetrics.registerFont(TTFont("NotoSansKR", str(NOTO_FONT)))
        return {"medium": "NotoSansKR", "bold": "NotoSansKR", "extra_bold": "NotoSansKR", "chart_path": str(NOTO_FONT)}
    if NOTO_VAR_FONT.exists():
        pdfmetrics.registerFont(TTFont("NotoSansKR", str(NOTO_VAR_FONT)))
        return {"medium": "NotoSansKR", "bold": "NotoSansKR", "extra_bold": "NotoSansKR", "chart_path": str(NOTO_VAR_FONT)}
    pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
    return {"medium": "HYGothic-Medium", "bold": "HYGothic-Medium", "extra_bold": "HYGothic-Medium", "chart_path": None}


def describe_active_report_font() -> str:
    fonts = register_pdf_fonts()
    if fonts["medium"].startswith("Dongle"):
        return "Dongle 둥근 폰트 적용됨: 본문 Regular, 제목/KPI Bold"
    if fonts["medium"].startswith("Pretendard"):
        return "Pretendard 적용됨: 본문 Medium, 제목 Bold, KPI 숫자 ExtraBold"
    if fonts["medium"] == "NotoSansKR":
        return "Noto Sans KR 적용 중: assets/fonts에 Pretendard 파일이 없으면 이 폰트로 대체됩니다."
    return f"{fonts['medium']} 적용 중"


def find_korean_ttf_font() -> str | None:
    pretendard_medium = find_pretendard_font("Medium") or find_pretendard_font("Regular")
    if pretendard_medium:
        return str(pretendard_medium)
    dongle_regular = find_dongle_font("Regular")
    if dongle_regular:
        return str(dongle_regular)
    for path in [
        NOTO_FONT,
        NOTO_VAR_FONT,
        Path("/System/Library/Fonts/Supplemental/AppleGothic.ttf"),
        Path("/Library/Fonts/AppleGothic.ttf"),
    ]:
        if path.exists():
            return str(path)
    return None


def make_styles(fonts: ReportFonts) -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("TitleKR", parent=base["Title"], fontName=fonts["extra_bold"], fontSize=20, leading=25, textColor=TEXT, alignment=TA_LEFT),
        "subtitle": ParagraphStyle("SubtitleKR", parent=base["Normal"], fontName=fonts["medium"], fontSize=10.5, leading=15, textColor=TEXT, alignment=TA_LEFT),
        "h2": ParagraphStyle("H2KR", parent=base["Heading2"], fontName=fonts["bold"], fontSize=10.5, leading=13, textColor=TEXT, spaceAfter=0),
        "body": ParagraphStyle("BodyKR", parent=base["BodyText"], fontName=fonts["medium"], fontSize=8.5, leading=12.5, textColor=TEXT, alignment=TA_LEFT),
        "small": ParagraphStyle("SmallKR", parent=base["BodyText"], fontName=fonts["medium"], fontSize=6.5, leading=8.2, textColor=MUTED, alignment=TA_CENTER),
        "card_label": ParagraphStyle("CardLabel", parent=base["BodyText"], fontName=fonts["medium"], fontSize=7.2, leading=9, textColor=MUTED, alignment=TA_CENTER),
        "card_value": ParagraphStyle("CardValue", parent=base["BodyText"], fontName=fonts["extra_bold"], fontSize=10.8, leading=13, textColor=TEXT, alignment=TA_CENTER),
        "chart_label": ParagraphStyle("ChartLabel", parent=base["BodyText"], fontName=fonts["bold"], fontSize=7.8, leading=8.8, textColor=TEXT, alignment=TA_CENTER),
        "notice": ParagraphStyle("NoticeKR", parent=base["BodyText"], fontName=fonts["medium"], fontSize=7.2, leading=9, textColor=MUTED, alignment=TA_LEFT),
    }


def draw_report_logo(canvas, width: float, height: float, page_number: int, inputs: StoreInputs) -> None:
    if page_number != 1:
        return
    logo_bytes = inputs.logo_bytes
    if logo_bytes is None:
        return
    try:
        with PILImage.open(io.BytesIO(logo_bytes)) as image:
            logo_width_px, logo_height_px = image.size
    except Exception:
        return
    if logo_width_px <= 0 or logo_height_px <= 0:
        return

    logo_width = 54 * mm
    logo_height = logo_width * (logo_height_px / logo_width_px)
    x = width - 13 * mm - logo_width
    y = height - 22 * mm
    canvas.drawImage(ImageReader(io.BytesIO(logo_bytes)), x, y, width=logo_width, height=logo_height, preserveAspectRatio=True, mask="auto")


def draw_page_background(canvas, doc, inputs: StoreInputs, fonts: ReportFonts) -> None:
    canvas.saveState()
    width, height = A4
    canvas.setFillColor(LIGHT_BG)
    canvas.rect(0, 0, width, height, fill=1, stroke=0)
    if doc.page == 1:
        header_h = 34 * mm
        canvas.setFillColor(ACCENT)
        canvas.rect(0, height - header_h, width, header_h, fill=1, stroke=0)
    draw_report_logo(canvas, width, height, doc.page, inputs)
    canvas.restoreState()


def paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(str(text).replace("\n", "<br/>"), style)


def section_heading(title: str, styles: dict[str, ParagraphStyle]) -> Table:
    title_row = Table([[paragraph(title, styles["h2"])]], colWidths=[176 * mm], hAlign="LEFT")
    title_row.setStyle(
        TableStyle(
            [
                ("LINEBEFORE", (0, 0), (0, 0), 3, ACCENT),
                ("LEFTPADDING", (0, 0), (0, 0), 6),
                ("BOTTOMPADDING", (0, 0), (0, 0), 2.5),
            ]
        )
    )
    line = Table([[""]], colWidths=[176 * mm], rowHeights=[1])
    line.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), LINE), ("FONTNAME", (0, 0), (-1, -1), styles["body"].fontName)]))
    return Table([[title_row], [line]], colWidths=[176 * mm], hAlign="LEFT", spaceBefore=3, spaceAfter=2.2)


def metric_cards(items: list[tuple[str, str]], styles: dict[str, ParagraphStyle]) -> Table:
    card_width = (176 * mm - 12 * mm) / 5
    row = []
    for idx, (label, value) in enumerate(items):
        if idx:
            row.append("")
        card = Table(
            [[paragraph(label, styles["card_label"])], [paragraph(value, styles["card_value"])]],
            colWidths=[card_width],
        )
        card.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                    ("BOX", (0, 0), (-1, -1), 0.35, LINE),
                    ("LINEABOVE", (0, 0), (-1, 0), 2, ACCENT),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 5.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5.5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        row.append(card)
    table = Table([row], colWidths=[card_width, 3 * mm, card_width, 3 * mm, card_width, 3 * mm, card_width, 3 * mm, card_width], hAlign="LEFT")
    table.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), styles["body"].fontName), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
    return table


def simple_table(data: list[list[object]], fonts: ReportFonts, widths: list[float], font_size: float = 7.6, row_padding: float = 3.2) -> Table:
    table = Table(data, colWidths=widths, hAlign="LEFT", repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), fonts["medium"]),
                ("FONTNAME", (0, 0), (-1, 0), fonts["bold"]),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("LEADING", (0, 0), (-1, -1), font_size + 2),
                ("BACKGROUND", (0, 0), (-1, 0), ACCENT_SOFT),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#2B2B2B")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.22, LINE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), row_padding),
                ("BOTTOMPADDING", (0, 0), (-1, -1), row_padding),
            ]
        )
    )
    return table


def comparison_info_table(current_df: pd.DataFrame, previous_df: pd.DataFrame, fonts: ReportFonts) -> Table:
    data = [
        ["비교 기간 :", f"{format_month_full(previous_df['날짜'].min())} ↔ {format_month_full(current_df['날짜'].min())}"],
        ["광고 운영일수 :", f"전월 {len(previous_df)}일 | 대상월 {len(current_df)}일"],
    ]
    table = Table(data, colWidths=[36 * mm, 140 * mm], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), fonts["medium"]),
                ("FONTNAME", (0, 0), (0, -1), fonts["bold"]),
                ("FONTSIZE", (0, 0), (-1, -1), 7.3),
                ("LEADING", (0, 0), (-1, -1), 9.3),
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#555555")),
                ("BOX", (0, 0), (-1, -1), 0.25, LINE),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, LINE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 2.0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.0),
            ]
        )
    )
    return table


def comparison_table(rows: list[dict[str, object]], fonts: ReportFonts, include_change: bool) -> Table:
    data = [["항목", "대상월", "전월", "전월 대비"]]
    for row in rows:
        data.append([row["label"], row["current_text"], row["previous_text"], row["change_text"]])
    col_widths = [42 * mm, 44 * mm, 44 * mm, 46 * mm]
    table = simple_table(data, fonts, col_widths, 7.1, 2.15)
    commands = [("ALIGN", (1, 0), (-1, -1), "RIGHT")]
    for idx, row in enumerate(rows, start=1):
        if include_change:
            current_value = float(row["current"])
            previous_value = float(row["previous"])
            color = MUTED if current_value == previous_value else colors.HexColor("#DC2626") if current_value > previous_value else colors.HexColor("#2563EB")
            commands.append(("TEXTCOLOR", (3, idx), (3, idx), color))
        else:
            commands.append(("TEXTCOLOR", (3, idx), (3, idx), MUTED))
    table.setStyle(TableStyle(commands))
    return table


def configure_matplotlib_font() -> None:
    font_path = find_korean_ttf_font()
    if font_path:
        from matplotlib import font_manager

        font_manager.fontManager.addfont(font_path)
        plt.rcParams["font.family"] = font_manager.FontProperties(fname=font_path).get_name()
    plt.rcParams["axes.unicode_minus"] = False


def chart_image(df: pd.DataFrame, column: str, color: str) -> io.BytesIO:
    configure_matplotlib_font()
    plot_df = df.copy()
    plot_df["표기"] = plot_df["날짜"].dt.strftime("%m.%d")
    plot_df["x"] = range(len(plot_df))
    tick_positions = list(plot_df["x"])
    tick_labels = [plot_df["표기"].iloc[idx] for idx in tick_positions]

    fig, ax = plt.subplots(figsize=(11.6, 3.75), dpi=220)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")
    ax.plot(plot_df["x"], plot_df[column], color=color, linewidth=2.1, marker="o", markersize=2.8)
    ax.fill_between(plot_df["x"], plot_df[column], color=color, alpha=0.12)
    ax.grid(True, color="#E8E4D8", alpha=0.55, linewidth=0.55)
    for spine in ax.spines.values():
        spine.set_color("#D8D3C7")
        spine.set_linewidth(0.7)
    ax.set_xticks(tick_positions)
    ax.set_xticklabels(tick_labels, fontsize=6.2, rotation=45, ha="right")
    ax.tick_params(axis="y", labelsize=8.5)
    ax.tick_params(colors="#333333")
    ax.margins(x=0.015)
    fig.subplots_adjust(left=0.07, right=0.99, top=0.96, bottom=0.27)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.06)
    plt.close(fig)
    buf.seek(0)
    return buf


def daily_chart_table(df: pd.DataFrame, styles: dict[str, ParagraphStyle]) -> Table:
    chart_width = 176 * mm
    chart_height = 57 * mm
    rows = []
    for idx, (title, column, color) in enumerate([("일별 노출수 추이", "노출수", "#111111"), ("일별 클릭수 추이", "클릭수", "#F2B705")]):
        if idx:
            rows.append([Spacer(1, 1 * mm)])
        rows.append([paragraph(title, styles["chart_label"])])
        rows.append([Spacer(1, 0.4 * mm)])
        rows.append([RLImage(chart_image(df, column, color), width=chart_width, height=chart_height)])
    table = Table(rows, colWidths=[chart_width], hAlign="LEFT")
    table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
    return table


def build_pdf(df: pd.DataFrame, inputs: StoreInputs, mapping: dict[str, str], previous_df: pd.DataFrame | None = None) -> bytes:
    fonts = register_pdf_fonts()
    styles = make_styles(fonts)
    metrics = calculate_metrics(df)
    previous_metrics = calculate_metrics(previous_df) if previous_df is not None and not previous_df.empty else None
    current_days = len(df)
    previous_days = len(previous_df) if previous_df is not None and not previous_df.empty else 0
    has_previous = previous_metrics is not None
    comparison_excluded = has_previous and abs(current_days - previous_days) >= 6
    include_change = has_previous and not comparison_excluded
    comparison_rows = build_month_comparison(
        metrics,
        previous_metrics,
        current_days=current_days,
        previous_days=previous_days or 1,
        include_change=include_change,
    )
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=13 * mm, leftMargin=13 * mm, topMargin=9 * mm, bottomMargin=9 * mm)
    story = []

    # Page 1: store, period, key metrics, comparison, charts.
    story.append(Spacer(1, 1 * mm))
    story.append(paragraph(inputs.store_name, styles["title"]))
    story.append(paragraph(f"{format_report_title_month(inputs.period_start)} CPC 운영 리포트", styles["subtitle"]))
    story.append(paragraph(f"운영기간: {format_period(inputs.period_start, inputs.period_end)}", styles["subtitle"]))
    story.append(Spacer(1, 4 * mm))

    story.append(section_heading("핵심 성과 지표", styles))
    story.append(
        metric_cards(
            [
                ("노출수", f"{format_number(metrics['total_impressions'])}회"),
                ("클릭수", f"{format_number(metrics['total_clicks'])}회"),
                ("광고비", format_currency(metrics["total_spend"])),
                ("클릭당 평균비용", format_cpc_currency(metrics["avg_cpc"])),
                ("상품권 조회수", f"{format_number(metrics['total_voucher'])}회"),
            ],
            styles,
        )
    )
    if has_previous:
        story.append(section_heading("전월 대비 성과 비교", styles))
        story.append(comparison_info_table(df, previous_df, fonts))
        story.append(Spacer(1, 1 * mm))
        story.append(comparison_table(comparison_rows, fonts, include_change))
        story.append(Spacer(1, 0.5 * mm))
        if comparison_excluded:
            story.append(paragraph("※ 집행 기간이 짧아 전월 대비 증감률은 제공되지 않으며, 전월 데이터는 참고용으로 제공됩니다.", styles["notice"]))
        elif include_change:
            story.append(paragraph("※ 전월 데이터 및 증감률은 참고용으로 제공됩니다.", styles["notice"]))
        else:
            story.append(paragraph("※ 전월 데이터는 참고용으로 제공되며, 증감률은 제공되지 않습니다.", styles["notice"]))

    story.append(section_heading("일별 성과 추이", styles))
    if not df.empty:
        story.append(daily_chart_table(df, styles))
    else:
        story.append(paragraph("그래프를 생성할 수 있는 일별 데이터가 충분하지 않습니다.", styles["body"]))

    story.append(PageBreak())

    # Page 2: daily table and owner-facing monthly summary.
    story.append(section_heading("일별 상세 데이터", styles))
    show_voucher_column = bool(df["상품권 조회수"].sum() > 0) if "상품권 조회수" in df.columns else False
    daily_headers = ["날짜", "광고비", "노출수", "클릭수", "클릭당 비용"]
    if show_voucher_column:
        daily_headers.append("상품권 조회수")
    daily_rows = [daily_headers]
    for _, row in df.head(31).iterrows():
        clicks = float(row["클릭수"])
        spend = float(row[SPEND_COL])
        daily_cpc = spend / clicks if clicks > 0 else None
        daily_line = [
            format_short_date(row["날짜"]),
            format_currency(spend),
            format_number(row["노출수"]),
            format_number(clicks),
            format_daily_cpc(daily_cpc),
        ]
        if show_voucher_column:
            daily_line.append(f"{format_number(row['상품권 조회수'])}회")
        daily_rows.append(daily_line)
    daily_widths = [22 * mm, 30 * mm, 28 * mm, 24 * mm, 32 * mm, 40 * mm] if show_voucher_column else [27 * mm, 37 * mm, 34 * mm, 31 * mm, 47 * mm]
    story.append(simple_table(daily_rows, fonts, daily_widths, 8.1, 3.6))
    if len(df) > 31:
        story.append(paragraph("일별 데이터가 많아 PDF 표에는 앞의 31개 행을 표시했습니다.", styles["body"]))

    def on_page(canvas, doc_obj):
        draw_page_background(canvas, doc_obj, inputs, fonts)
        canvas.saveState()
        canvas.setFont(fonts["medium"], 8)
        canvas.setFillColor(MUTED)
        canvas.drawRightString(A4[0] - 16 * mm, 8 * mm, f"{doc_obj.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    return buffer.getvalue()


def make_file_name(inputs: StoreInputs) -> str:
    return f"{inputs.store_name}_{inputs.period_start.year}년{inputs.period_start.month:02d}월_CPC운영보고서.pdf"


def make_image_file_name(inputs: StoreInputs) -> str:
    return f"{inputs.store_name}_{inputs.period_start.year}년{inputs.period_start.month:02d}월_CPC운영보고서.png"


def crop_page_for_long_png(image: PILImage.Image) -> PILImage.Image:
    rgb = image.convert("RGB")
    width, height = rgb.size
    px_per_mm = height / 297
    scan_bottom = max(0, height - int(13 * px_per_mm))
    background = (250, 250, 247)
    threshold = 34
    y_min = height
    y_max = 0
    pixels = rgb.load()

    for y in range(scan_bottom):
        row_has_content = False
        for x in range(width):
            r, g, b = pixels[x, y]
            if abs(r - background[0]) + abs(g - background[1]) + abs(b - background[2]) > threshold:
                row_has_content = True
                break
        if row_has_content:
            y_min = min(y_min, y)
            y_max = max(y_max, y)

    if y_max <= y_min:
        return image

    padding = int(5 * px_per_mm)
    top = max(0, y_min - padding)
    bottom = min(height, y_max + padding)
    return image.crop((0, top, width, bottom))


def pdf_to_png_bytes(pdf_bytes: bytes, zoom: float = 2.0) -> bytes:
    if fitz is None:
        raise RuntimeError("이미지 생성을 위해 PyMuPDF 패키지가 필요합니다. requirements.txt 설치 후 다시 실행해주세요.")
    document = fitz.open(stream=pdf_bytes, filetype="pdf")
    images: list[PILImage.Image] = []
    matrix = fitz.Matrix(zoom, zoom)
    for page in document:
        pix = page.get_pixmap(matrix=matrix, alpha=False)
        page_image = PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
        images.append(crop_page_for_long_png(page_image))
    gap = 0
    width = max(image.width for image in images)
    height = sum(image.height for image in images) + gap * (len(images) - 1)
    combined = PILImage.new("RGB", (width, height), "white")
    y = 0
    for image in images:
        combined.paste(image, ((width - image.width) // 2, y))
        y += image.height + gap
    output = io.BytesIO()
    combined.save(output, format="PNG", optimize=True)
    return output.getvalue()


ROMAN_INITIAL = ["g", "kk", "n", "d", "tt", "r", "m", "b", "pp", "s", "ss", "", "j", "jj", "ch", "k", "t", "p", "h"]
ROMAN_MEDIAL = ["a", "ae", "ya", "yae", "eo", "e", "yeo", "ye", "o", "wa", "wae", "oe", "yo", "u", "wo", "we", "wi", "yu", "eu", "ui", "i"]
ROMAN_FINAL = ["", "k", "k", "ks", "n", "nj", "nh", "t", "l", "lk", "lm", "lb", "ls", "lt", "lp", "lh", "m", "p", "ps", "t", "t", "ng", "t", "t", "k", "t", "p", "t"]


def romanize_korean_text(text: str) -> str:
    parts: list[str] = []
    for char in text:
        code = ord(char)
        if 0xAC00 <= code <= 0xD7A3:
            syllable = code - 0xAC00
            initial = syllable // 588
            medial = (syllable % 588) // 28
            final = syllable % 28
            parts.append(ROMAN_INITIAL[initial] + ROMAN_MEDIAL[medial] + ROMAN_FINAL[final])
        else:
            parts.append(char)
    text = "".join(parts)
    text = re.sub(r"\s+", " ", text).strip()
    return text.title()


def auto_english_address(address_kr: str) -> str:
    if not address_kr.strip():
        return ""
    if address_kr.isascii():
        return address_kr.strip()
    return romanize_korean_text(address_kr)


def passport_style_name(name_kr: str) -> str:
    return romanize_korean_text(name_kr).upper()


def normalize_ocr_line(text: str) -> str:
    return re.sub(r"\s+", " ", text.replace("|", " ").replace(":", " ")).strip()


def parse_business_registration_text(text: str) -> dict[str, str]:
    lines = [normalize_ocr_line(line) for line in text.splitlines() if normalize_ocr_line(line)]
    joined = "\n".join(lines)

    def find_after(labels: list[str]) -> str:
        for label in labels:
            pattern = rf"{label}\s*[:：]?\s*([^\n]+)"
            match = re.search(pattern, joined)
            if match:
                return match.group(1).strip()
        return ""

    business_address = find_after(["사업장소재지", "사업장 소재지", "사업장주소", "사업장 주소"])
    headquarters_address = find_after(["본점소재지", "본점 소재지", "본점주소", "본점 주소"])
    address = business_address or headquarters_address
    owner_name = find_after(["대표자명", "대표자", "성명", "법인대표자명", "법인 대표자명"])
    business_name = find_after(["상호", "법인명", "사업자명", "단체명"])

    return {
        "business_address": business_address,
        "headquarters_address": headquarters_address,
        "selected_address": address,
        "owner_name": owner_name,
        "business_name": business_name,
    }


def ocr_business_registration(image_bytes: bytes) -> tuple[dict[str, str], list[str]]:
    messages: list[str] = []
    try:
        import pytesseract
    except ImportError:
        messages.append("OCR 패키지가 설치되어 있지 않아 사업자등록증 자동 인식을 건너뛰었습니다. 아래 입력값은 직접 수정할 수 있습니다.")
        return {}, messages

    try:
        image = PILImage.open(io.BytesIO(image_bytes))
        text = pytesseract.image_to_string(image, lang="kor+eng")
    except Exception as exc:
        messages.append(f"OCR 처리에 실패했습니다. 직접 입력값을 확인해주세요. 오류: {exc}")
        return {}, messages

    parsed = parse_business_registration_text(text)
    if parsed.get("business_address") and parsed.get("headquarters_address"):
        if parsed["business_address"] != parsed["headquarters_address"]:
            messages.append("본점소재지와 사업장소재지가 달라 사업장소재지를 우선 사용했습니다.")
        else:
            messages.append("본점소재지와 사업장소재지가 동일하여 해당 주소를 사용했습니다.")
    elif parsed.get("selected_address"):
        messages.append("사업자등록증에서 주소를 인식했습니다.")
    else:
        messages.append("사업자등록증에서 주소를 자동 인식하지 못했습니다.")
    if parsed.get("owner_name"):
        messages.append("대표자명/법인대표자명을 자동 인식했습니다.")
    return parsed, messages


def safe_filename(value: str) -> str:
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value.strip())
    return value or "계약서"


def merged_anchor(ws, cell: str):
    target = ws[cell]
    for merged_range in ws.merged_cells.ranges:
        if target.coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col)
    return target


def set_cell_value(ws, cell: str, value: object) -> None:
    target = merged_anchor(ws, cell)
    target.value = value
    target.alignment = target.alignment.copy(wrap_text=True, vertical="center")


def clear_yellow_input_fills(ws) -> None:
    from openpyxl.styles import PatternFill

    no_fill = PatternFill(fill_type=None)
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            rgb = getattr(fill.fgColor, "rgb", None)
            indexed = getattr(fill.fgColor, "indexed", None)
            if rgb in YELLOW_FILL_RGB or indexed == 13:
                cell.fill = no_fill


def add_image_to_sheet(ws, image_bytes: bytes | None, anchor: str, width: int, height: int) -> None:
    if not image_bytes:
        return
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from PIL import Image as PILImage

    with PILImage.open(io.BytesIO(image_bytes)) as source_image:
        original_width, original_height = source_image.size
        scale = min(width / original_width, height / original_height) if original_width and original_height else 1

    with NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp.write(image_bytes)
        tmp_path = tmp.name
    image = OpenpyxlImage(tmp_path)
    image.width = max(1, int(original_width * scale))
    image.height = max(1, int(original_height * scale))
    ws.add_image(image, anchor)


def make_contract_file_name(inputs: ContractInputs, ext: str) -> str:
    return f"{safe_filename(inputs.store_name)}_계약서.{ext}"


def fill_contract_workbook(inputs: ContractInputs) -> bytes:
    from openpyxl import load_workbook

    template_path = CONTRACT_TEMPLATES[inputs.contract_type]
    if not template_path.exists():
        raise FileNotFoundError(f"{inputs.contract_type} 계약서 템플릿을 찾을 수 없습니다: {template_path}")

    workbook = load_workbook(template_path)
    ws = workbook["中文版"] if "中文版" in workbook.sheetnames else workbook.active

    start_text = f"{inputs.contract_start.year}年 {inputs.contract_start.month:02d}月{inputs.contract_start.day:02d}日"
    duration_text = "自开通之日起【365】天" if inputs.contract_type == "1년" else "自开通之日起【3】个月"
    kadob_suffix = re.sub(r"^NO\.?\s*KADOB", "", inputs.kadob_no.strip(), flags=re.IGNORECASE)
    kadob_text = f"NO.KADOB{kadob_suffix}" if kadob_suffix else "NO.KADOB"

    set_cell_value(ws, "M1", kadob_text)
    set_cell_value(ws, "E4", inputs.store_name)
    set_cell_value(ws, "J4", inputs.address_en)
    set_cell_value(ws, "E5", inputs.passport_english_name or inputs.business_name_kr)
    set_cell_value(ws, "L5", inputs.passport_number)
    set_cell_value(ws, "D14", inputs.store_name)
    set_cell_value(ws, "H14", inputs.region)
    set_cell_value(ws, "B19", f"{inputs.store_name} / {inputs.store_code}".strip(" /"))
    set_cell_value(ws, "D19", inputs.region)
    set_cell_value(ws, "C12", f"☐ 推广计划开始时间 {start_text}")
    set_cell_value(ws, "L19", duration_text)
    set_cell_value(ws, "B61", f"签字日期： {start_text}")
    if inputs.memo.strip():
        set_cell_value(ws, "C54", f"{ws['C54'].value}\n补充备注：{inputs.memo.strip()}")

    add_image_to_sheet(ws, inputs.signature_bytes, "D58", 110, 38)
    add_image_to_sheet(ws, inputs.stamp_bytes, "F57", 72, 72)
    clear_yellow_input_fills(ws)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def contract_pdf_bytes(inputs: ContractInputs) -> bytes:
    fonts = register_pdf_fonts()
    body_font = fonts["medium"]
    styles = getSampleStyleSheet()
    title = ParagraphStyle("ContractTitle", parent=styles["Title"], fontName=fonts["extra_bold"], fontSize=18, leading=23, textColor=TEXT)
    h2 = ParagraphStyle("ContractH2", parent=styles["Heading2"], fontName=fonts["bold"], fontSize=11, leading=14, textColor=TEXT)
    body = ParagraphStyle("ContractBody", parent=styles["BodyText"], fontName=body_font, fontSize=8.5, leading=12, textColor=TEXT)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    rows = [
        ["계약기간", CONTRACT_TYPE_LABELS.get(inputs.contract_type, inputs.contract_type)],
        ["NO.KADOB", inputs.kadob_no or "-"],
        ["매장명", inputs.store_name],
        ["영문 주소", inputs.address_en],
        ["사업자명", inputs.business_name_kr or "-"],
        ["여권 영문명", inputs.passport_english_name or "-"],
        ["여권번호", inputs.passport_number or "-"],
        ["지역", inputs.region],
        ["매장코드", inputs.store_code],
        ["계약/운영 시작일", inputs.contract_start.strftime("%Y.%m.%d")],
    ]
    story = [
        Paragraph("계약서", title),
        Spacer(1, 4 * mm),
        Paragraph("계약서 핵심 입력 정보", h2),
    ]
    table = Table(rows, colWidths=[45 * mm, 135 * mm], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), body_font),
                ("FONTNAME", (0, 0), (0, -1), fonts["bold"]),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("BACKGROUND", (0, 0), (0, -1), ACCENT_SOFT),
                ("GRID", (0, 0), (-1, -1), 0.25, LINE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("※ 정식 계약 양식은 함께 생성되는 Excel 파일에 원본 중국어 템플릿 형식으로 반영됩니다.", body))
    if inputs.memo.strip():
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(f"비고: {inputs.memo.strip()}", body))
    if inputs.signature_bytes:
        story.append(Spacer(1, 8 * mm))
        story.append(Paragraph("서명", h2))
        story.append(RLImage(io.BytesIO(inputs.signature_bytes), width=48 * mm, height=18 * mm))
    if inputs.stamp_bytes:
        story.append(Spacer(1, 8 * mm))
        story.append(Paragraph("도장", h2))
        story.append(RLImage(io.BytesIO(inputs.stamp_bytes), width=38 * mm, height=22 * mm))
    doc.build(story)
    return buffer.getvalue()


def contract_zip_bytes(inputs: ContractInputs, xlsx_bytes: bytes, pdf_bytes: bytes) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(make_contract_file_name(inputs, "xlsx"), xlsx_bytes)
        archive.writestr(make_contract_file_name(inputs, "pdf"), pdf_bytes)
    output.seek(0)
    return output.getvalue()


def preview_metrics(df: pd.DataFrame) -> None:
    metrics = calculate_metrics(df)
    items = [
        ("노출수", f"{format_number(metrics['total_impressions'])}회"),
        ("클릭수", f"{format_number(metrics['total_clicks'])}회"),
        ("광고비", format_currency(metrics["total_spend"])),
        ("클릭당 평균비용(CPC)", format_cpc_currency(metrics["avg_cpc"])),
        ("상품권 조회수", f"{format_number(metrics['total_voucher'])}회"),
    ]
    cols = st.columns(len(items))
    for col, (label, value) in zip(cols, items):
        col.metric(label, value)


def run_cpc_report_app() -> None:
    left, right = st.columns([1, 1.1], gap="large")
    with left:
        with st.container(border=True):
            uploaded_files = st.file_uploader("CPC 데이터 파일 업로드", type=["xlsx", "xls", "csv"], accept_multiple_files=True)
            use_logo = st.checkbox("보고서에 로고 표시", value=False)
            uploaded_logo = None
            if use_logo:
                uploaded_logo = st.file_uploader("로고 이미지 선택", type=["png", "jpg", "jpeg"], help="선택하지 않으면 기본 로고가 사용됩니다.")
            store_name = st.text_input("매장명", placeholder="예: 홍대점")
            use_pause = st.checkbox("광고 중단일 입력")
            pause_date = st.date_input("광고 중단일", value=date.today()) if use_pause else None
            memo = st.text_area("비고", placeholder="선택 입력")
            generate = st.button("보고서 생성", type="primary", use_container_width=True)

    with right:
        with st.container(border=True):
            st.subheader("생성 안내")
            st.markdown(
                """
                <div class="hint">
                당월 파일과 전월 파일을 한 번에 함께 업로드할 수 있습니다. 업로드된 파일의 날짜 데이터를 기준으로 가장 최근 월을 당월, 직전 월을 전월 비교 데이터로 자동 인식합니다.
                </div>
                """,
                unsafe_allow_html=True,
            )

    if not generate:
        return
    if not uploaded_files:
        st.error("CPC 엑셀 또는 CSV 파일을 먼저 업로드해주세요.")
        return
    if not store_name.strip():
        st.error("매장명을 입력해주세요.")
        return
    try:
        with st.spinner("데이터를 정리하고 2페이지 보고서를 만드는 중입니다..."):
            processed_files = []
            mapping: dict[str, str] = {}
            messages: list[str] = []
            month_messages: list[str] = []
            for uploaded_file in uploaded_files:
                raw_df = read_best_sheet(uploaded_file)
                file_df, file_mapping, file_messages = preprocess_dataframe(raw_df)
                mapping.update(file_mapping)
                messages.extend([f"{uploaded_file.name}: {message}" for message in file_messages])
                if file_df.empty:
                    month_messages.append(f"{uploaded_file.name}: 날짜가 포함된 데이터 행을 찾지 못해 제외했습니다.")
                    continue
                month = latest_month_period(file_df)
                if month is None:
                    month_messages.append(f"{uploaded_file.name}: 월별 데이터를 확인하지 못해 제외했습니다.")
                    continue
                processed_files.append({"name": uploaded_file.name, "df": file_df, "month": month})

            if not processed_files:
                st.warning("날짜가 포함된 데이터 행을 찾지 못했습니다. 업로드 파일의 날짜 컬럼을 확인해주세요.")
                return

            processed_files = sorted(processed_files, key=lambda item: item["month"])
            latest_month = processed_files[-1]["month"]
            current_sources = [item for item in processed_files if item["month"] == latest_month]
            current_all_df = pd.concat([item["df"] for item in current_sources], ignore_index=True)
            df, auto_previous_df, period_start, period_end, split_messages = split_latest_months(current_all_df)
            month_messages.extend(split_messages)
            month_messages.append(f"당월 데이터: {', '.join(item['name'] for item in current_sources)}")

            previous_month = latest_month - 1
            previous_sources = [item for item in processed_files if item["month"] == previous_month]
            if previous_sources:
                previous_all_df = pd.concat([item["df"] for item in previous_sources], ignore_index=True)
                previous_df, _, _, previous_month_messages = select_latest_month(previous_all_df, "전월")
                month_messages.extend(previous_month_messages)
                month_messages.append(f"전월 데이터: {', '.join(item['name'] for item in previous_sources)}")
            else:
                previous_df = auto_previous_df

            logo_bytes = None
            if use_logo:
                if uploaded_logo is not None:
                    logo_bytes = uploaded_logo.getvalue()
                elif LOGO_PATH.exists():
                    logo_bytes = LOGO_PATH.read_bytes()
            inputs = StoreInputs(store_name.strip(), period_start, period_end, pause_date, memo, logo_bytes)
            pdf_bytes = build_pdf(df, inputs, mapping, previous_df)
            png_bytes = pdf_to_png_bytes(pdf_bytes)

        st.success("PDF와 이미지 보고서가 생성되었습니다.")
        st.caption(f"현재 보고서 폰트: {describe_active_report_font()}")
        preview_metrics(df)
        if messages or month_messages:
            with st.expander("자동 처리 안내"):
                for message in messages + month_messages:
                    st.info(message)
        cols = st.columns(2)
        cols[0].download_button("PDF 다운로드", data=pdf_bytes, file_name=make_file_name(inputs), mime="application/pdf", use_container_width=True)
        cols[1].download_button("이미지(PNG) 다운로드", data=png_bytes, file_name=make_image_file_name(inputs), mime="image/png", use_container_width=True)
        with st.expander("일별 데이터 미리보기"):
            preview = df.copy()
            preview["날짜"] = preview["날짜"].dt.strftime("%m.%d")
            preview_columns = ["날짜", SPEND_COL, "노출수", "클릭수", "상품권 조회수", "저장하기", "공유하기"]
            st.dataframe(preview[[col for col in preview_columns if col in preview.columns]], use_container_width=True)
    except Exception as exc:
        st.error("보고서를 생성하는 중 문제가 발생했습니다.")
        st.info(f"확인할 내용: 파일 형식, 날짜 컬럼, 숫자 컬럼 값을 확인해주세요. 오류 내용: {exc}")


def run_contract_app() -> None:
    missing_templates = [label for label, path in CONTRACT_TEMPLATES.items() if not path.exists()]
    if missing_templates:
        st.error(f"계약서 템플릿 파일이 없습니다: {', '.join(missing_templates)}")
        return

    left, right = st.columns([1, 1], gap="large")
    with left:
        with st.container(border=True):
            st.subheader("계약서 정보 입력")
            contract_type_label = st.radio("계약기간", ["1年", "3个月"], horizontal=True)
            contract_type = "1년" if contract_type_label == "1年" else "3개월"
            kadob_no = st.text_input("NO.KADOB", key="contract_kadob_no", placeholder="예: 9030256")
            business_license = st.file_uploader("사업자등록증 이미지 업로드", type=["png", "jpg", "jpeg"], help="OCR 가능 환경에서는 주소와 대표자명을 자동 인식합니다.")

            ocr_messages: list[str] = []
            if business_license is not None:
                license_bytes = business_license.getvalue()
                upload_marker = f"{business_license.name}:{len(license_bytes)}"
                if st.session_state.get("contract_ocr_marker") != upload_marker:
                    parsed, ocr_messages = ocr_business_registration(license_bytes)
                    st.session_state["contract_ocr_marker"] = upload_marker
                    st.session_state["contract_ocr_messages"] = ocr_messages
                    if parsed.get("selected_address"):
                        st.session_state["contract_address_kr"] = parsed["selected_address"]
                        st.session_state["contract_address_en"] = auto_english_address(parsed["selected_address"])
                    if parsed.get("owner_name"):
                        st.session_state["contract_business_name_kr"] = parsed["owner_name"]
                        st.session_state["contract_passport_english_name"] = passport_style_name(parsed["owner_name"])
                    if parsed.get("business_name") and not st.session_state.get("contract_store_name"):
                        st.session_state["contract_store_name"] = parsed["business_name"]
                else:
                    ocr_messages = st.session_state.get("contract_ocr_messages", [])
            for message in ocr_messages:
                st.info(message)

            store_name = st.text_input("商户名称 / 门店1", key="contract_store_name", placeholder="예: 태초갈비 홍대점")
            address_kr = st.text_area("사업자등록증 주소", key="contract_address_kr", placeholder="본점소재지/사업장소재지 인식 또는 직접 입력")
            suggested_address = auto_english_address(address_kr)
            if address_kr and not st.session_state.get("contract_address_en"):
                st.session_state["contract_address_en"] = suggested_address
            address_en = st.text_area("商户地址 (영문)", key="contract_address_en", help="OCR/자동 변환값을 확인 후 필요하면 직접 수정해주세요.")
            business_name_kr = st.text_input("经营者/法定代表人姓名 원문", key="contract_business_name_kr", placeholder="예: 홍길동")
            if business_name_kr and not st.session_state.get("contract_passport_english_name"):
                st.session_state["contract_passport_english_name"] = passport_style_name(business_name_kr)
            passport_english_name = st.text_input("经营者/法定代表人姓名 (여권 영문명)", key="contract_passport_english_name", placeholder="예: HONG GILDONG")
            passport_number = st.text_input("身份证号码/护照号码", key="contract_passport_number", placeholder="선택 입력")
            region = st.selectbox("所在城市", CITY_OPTIONS, index=0)
            store_code = st.text_input("매장코드", placeholder="예: 12345678")
            contract_start = st.date_input("推广计划开始时间", value=date.today())

    with right:
        with st.container(border=True):
            st.subheader("서명 및 출력")
            st.info("계약서 템플릿에 있는 금액은 원본 그대로 유지합니다.")
            memo = st.text_area("비고", key="contract_memo", placeholder="선택 입력")
            signature_file = st.file_uploader("서명 이미지 업로드", type=["png", "jpg", "jpeg"], help="있으면 甲方 서명 영역에 삽입됩니다.")
            stamp_file = st.file_uploader("도장 이미지 업로드", type=["png", "jpg", "jpeg"], help="있으면 甲方 도장 영역에 삽입됩니다.")
            output_type = st.radio("출력 형식", ["Excel (.xlsx)", "PDF (.pdf)", "Excel + PDF"], horizontal=False)
            generate_contract = st.button("계약서 생성", type="primary", use_container_width=True)
            st.markdown(
                """
                <div class="hint">
                1年/3个月 선택에 따라 앱 내부 계약서 템플릿이 자동 적용됩니다. 노란색 입력 영역은 생성 시 값 입력 후 제거됩니다.
                </div>
                """,
                unsafe_allow_html=True,
            )

    if not generate_contract:
        return
    missing_fields = []
    if not kadob_no.strip():
        missing_fields.append("NO.KADOB")
    if not store_name.strip():
        missing_fields.append("商户名称")
    if not address_en.strip():
        missing_fields.append("商户地址")
    if not passport_english_name.strip() and not business_name_kr.strip():
        missing_fields.append("经营者/法定代表人姓名")
    if missing_fields:
        st.error(f"필수 항목이 누락되었습니다: {', '.join(missing_fields)}")
        return

    try:
        inputs = ContractInputs(
            contract_type=contract_type,
            kadob_no=kadob_no.strip(),
            store_name=store_name.strip(),
            address_kr=address_kr.strip(),
            address_en=address_en.strip() or suggested_address,
            business_name_kr=business_name_kr.strip(),
            passport_english_name=passport_english_name.strip(),
            passport_number=passport_number.strip(),
            region=region.strip(),
            store_code=store_code.strip(),
            contract_start=contract_start,
            cpc_amount=0.0,
            daily_budget=0.0,
            total_amount=0.0,
            memo=memo.strip(),
            signature_bytes=signature_file.getvalue() if signature_file is not None else None,
            stamp_bytes=stamp_file.getvalue() if stamp_file is not None else None,
        )
        with st.spinner("내부 템플릿에 계약 정보를 입력하는 중입니다..."):
            xlsx_bytes = fill_contract_workbook(inputs)
            pdf_bytes = contract_pdf_bytes(inputs)
        st.success("계약서 Excel/PDF가 생성되었습니다.")
        if output_type == "Excel (.xlsx)":
            st.download_button("계약서 Excel 다운로드", data=xlsx_bytes, file_name=make_contract_file_name(inputs, "xlsx"), mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        elif output_type == "PDF (.pdf)":
            st.download_button("계약서 PDF 다운로드", data=pdf_bytes, file_name=make_contract_file_name(inputs, "pdf"), mime="application/pdf", use_container_width=True)
        else:
            zip_bytes = contract_zip_bytes(inputs, xlsx_bytes, pdf_bytes)
            st.download_button("계약서 Excel + PDF 다운로드", data=zip_bytes, file_name=make_contract_file_name(inputs, "zip"), mime="application/zip", use_container_width=True)
    except Exception as exc:
        st.error("계약서를 생성하는 중 문제가 발생했습니다.")
        st.info(f"템플릿 파일, 입력값, 도장 PNG 파일을 확인해주세요. 오류 내용: {exc}")


def render_cpc_report_page() -> None:
    render_page_header(
        APP_TITLE,
        "여러 CPC 파일을 함께 업로드하면 가장 최근 월을 당월로 인식하고, 직전 월 데이터가 있으면 전월 비교까지 반영합니다.",
        "업무 자동화",
    )
    run_cpc_report_app()


def render_contract_page() -> None:
    render_page_header(
        "계약서 자동 생성",
        "사업자등록증 OCR, 계약 정보 입력, 서명/도장 이미지를 반영해 Excel/PDF/ZIP 계약서를 생성합니다.",
        "업무 자동화",
    )
    import contract_auto

    contract_auto.render_app(embedded=True)


def render_coming_soon_page(title: str, description: str, group: str) -> None:
    render_page_header(title, description, group)
    st.markdown(
        """
        <div class="coming-soon">
            <h2>준비 중</h2>
            <p>이 메뉴는 내부 업무 자동화 사이트 구조에 먼저 추가해두었습니다. 이후 매뉴얼 본문, 첨부 파일, 검색 색인 등을 연결할 수 있습니다.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_manual_page(item: NavItem) -> None:
    if item.key == "manual_search":
        render_page_header(item.label, "업무 매뉴얼 콘텐츠를 준비 중입니다.", item.group)
        search_term = st.text_input("매뉴얼 검색", placeholder="검색어를 입력하세요")
        if search_term:
            st.info("아직 등록된 매뉴얼 본문이 없어 검색 결과가 없습니다.")
        else:
            st.caption("매뉴얼 문서가 등록되면 제목과 본문 기준으로 검색 결과가 표시됩니다.")
        return
    render_coming_soon_page(item.label, "업무 매뉴얼 콘텐츠를 준비 중입니다.", item.group)


def render_page(selected_key: str) -> None:
    item = NAV_BY_KEY.get(selected_key, NAV_BY_KEY[DEFAULT_NAV_KEY])
    if item.key == "cpc_report":
        render_cpc_report_page()
        return
    if item.key == "contract_generator":
        render_contract_page()
        return
    render_manual_page(item)


def main() -> None:
    setup_page()
    selected_key = render_sidebar()
    render_page(selected_key)


if __name__ == "__main__":
    main()
