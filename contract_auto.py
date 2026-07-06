from __future__ import annotations

import io
import re
import zipfile
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import streamlit as st


BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "assets" / "templates"
CONTRACT_TEMPLATES = {
    "1年": TEMPLATE_DIR / "contract_1year.xlsx",
    "3个月": TEMPLATE_DIR / "contract_3months.xlsx",
}
CITY_OPTIONS = ["首尔", "济州岛", "仁川", "釜山", "京畿道"]
YELLOW_FILL_RGB = {"FFFFFF00", "FFFF00", "00FFFF00"}
PROVINCE_EN = {
    "서울특별시": "Seoul",
    "서울시": "Seoul",
    "부산광역시": "Busan",
    "부산시": "Busan",
    "인천광역시": "Incheon",
    "인천시": "Incheon",
    "제주특별자치도": "Jeju-do",
    "제주도": "Jeju-do",
    "경기도": "Gyeonggi-do",
    "강원도": "Gangwon-do",
    "충청북도": "Chungcheongbuk-do",
    "충청남도": "Chungcheongnam-do",
    "전라북도": "Jeollabuk-do",
    "전라남도": "Jeollanam-do",
    "경상북도": "Gyeongsangbuk-do",
    "경상남도": "Gyeongsangnam-do",
}


@dataclass
class ContractInputs:
    contract_type: str
    kadob_no: str
    store_name: str
    merchant_address: str
    owner_name: str
    city: str
    shop_id: str
    start_date: date
    ad_product: str
    memo: str
    signature_bytes: bytes | None = None
    stamp_bytes: bytes | None = None


def safe_filename(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", value.strip()) or "계약서"


def romanize_korean(text: str) -> str:
    initial = ["g", "kk", "n", "d", "tt", "r", "m", "b", "pp", "s", "ss", "", "j", "jj", "ch", "k", "t", "p", "h"]
    medial = ["a", "ae", "ya", "yae", "eo", "e", "yeo", "ye", "o", "wa", "wae", "oe", "yo", "u", "wo", "we", "wi", "yu", "eu", "ui", "i"]
    final = ["", "k", "k", "ks", "n", "nj", "nh", "t", "l", "lk", "lm", "lb", "ls", "lt", "lp", "lh", "m", "p", "ps", "t", "t", "ng", "t", "t", "k", "t", "p", "t"]
    parts = []
    for char in text:
        code = ord(char)
        if 0xAC00 <= code <= 0xD7A3:
            syllable = code - 0xAC00
            parts.append(initial[syllable // 588] + medial[(syllable % 588) // 28] + final[syllable % 28])
        else:
            parts.append(char)
    return re.sub(r"\s+", " ", "".join(parts)).strip().upper()


def romanize_token(token: str) -> str:
    suffixes = [
        ("시", "-si"),
        ("군", "-gun"),
        ("구", "-gu"),
        ("읍", "-eup"),
        ("면", "-myeon"),
        ("동", "-dong"),
        ("리", "-ri"),
        ("대로", "-daero"),
        ("로", "-ro"),
        ("길", "-gil"),
    ]
    for suffix, english_suffix in suffixes:
        if token.endswith(suffix) and len(token) > len(suffix):
            return romanize_korean(token[: -len(suffix)]).title().replace(" ", "") + english_suffix
    return romanize_korean(token).title().replace(" ", "")


def auto_english_address(address: str) -> str:
    address = re.sub(r"\([^)]*\)", "", address.strip())
    if not address:
        return ""
    if address.isascii():
        return address
    building_no = province = city = district = town = road = ""
    for token in address.split():
        if token in PROVINCE_EN:
            province = PROVINCE_EN[token]
        elif re.fullmatch(r"\d+(-\d+)?", token):
            building_no = token
        elif token.endswith(("로", "길", "대로")):
            road = romanize_token(token)
        elif token.endswith(("읍", "면", "동", "리")) and not town:
            town = romanize_token(token)
        elif token.endswith(("시", "군")) and not city:
            city = romanize_token(token)
        elif token.endswith("구") and not district:
            district = romanize_token(token)
    parts = [part for part in [building_no, road, town, district, city, province, "Republic of Korea"] if part]
    return ", ".join(parts) if len(parts) >= 3 else romanize_korean(address).title()


def parse_business_registration_text(text: str) -> dict[str, str]:
    lines = [re.sub(r"\s+", " ", line.replace(":", " ")).strip() for line in text.splitlines()]
    joined = "\n".join(line for line in lines if line)

    def find_after(labels: list[str]) -> str:
        for label in labels:
            match = re.search(rf"{label}\s*([^\n]+)", joined)
            if match:
                return match.group(1).strip()
        return ""

    business_address = find_after(["사업장소재지", "사업장 소재지", "사업장주소", "사업장 주소"])
    headquarters_address = find_after(["본점소재지", "본점 소재지", "본점주소", "본점 주소"])
    owner_name = find_after(["대표자명", "대표자", "성명", "법인대표자명", "법인 대표자명"])
    business_name = find_after(["상호", "법인명", "사업자명", "단체명"])
    return {
        "business_name": business_name,
        "owner_name": owner_name,
        "business_address": business_address or headquarters_address,
    }


def preprocess_image_basic(image):
    from PIL import ImageEnhance, ImageFilter, ImageOps

    image = ImageOps.exif_transpose(image).convert("RGB")
    longest = max(image.size)
    if longest < 1800:
        scale = 1800 / longest
        image = image.resize((int(image.width * scale), int(image.height * scale)))
    image = ImageOps.autocontrast(image)
    image = ImageEnhance.Contrast(image).enhance(1.35)
    image = ImageEnhance.Sharpness(image).enhance(1.25)
    return image.filter(ImageFilter.UnsharpMask(radius=1.2, percent=130, threshold=3))


def uploaded_file_to_images(file_bytes: bytes, filename: str):
    from PIL import Image as PILImage

    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        try:
            import fitz
        except Exception as exc:
            return [], [f"PDF OCR 패키지(PyMuPDF)가 없어 PDF 자동 인식을 건너뜁니다: {exc}"]
        try:
            document = fitz.open(stream=file_bytes, filetype="pdf")
            images = []
            for page_index in range(min(len(document), 3)):
                page = document.load_page(page_index)
                pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                images.append(PILImage.open(io.BytesIO(pixmap.tobytes("png"))))
            document.close()
            return images, [f"PDF {len(images)}페이지를 이미지로 변환했습니다."]
        except Exception as exc:
            return [], [f"PDF 변환 실패: {exc}"]
    if suffix in {".heic", ".heif"}:
        try:
            from pillow_heif import register_heif_opener

            register_heif_opener()
        except Exception as exc:
            return [], [f"HEIC 자동 인식 패키지가 없어 HEIC를 열 수 없습니다: {exc}"]
    try:
        return [PILImage.open(io.BytesIO(file_bytes))], []
    except Exception as exc:
        return [], [f"이미지 파일을 열 수 없습니다: {exc}"]


def ocr_business_registration(file_bytes: bytes, filename: str) -> tuple[dict[str, str], list[str], bytes | None]:
    try:
        import pytesseract
    except Exception as exc:
        return {}, [f"OCR 패키지가 없어 자동 인식을 건너뜁니다: {exc}", "아래 입력칸에 직접 입력해 주세요."], None

    images, messages = uploaded_file_to_images(file_bytes, filename)
    if not images:
        return {}, messages, None
    processed_images = [preprocess_image_basic(image) for image in images]
    preview = io.BytesIO()
    processed_images[0].save(preview, format="PNG")
    text_parts = []
    for index, image in enumerate(processed_images, 1):
        try:
            text_parts.append(pytesseract.image_to_string(image, lang="kor+eng", config="--psm 6"))
        except Exception as exc:
            messages.append(f"{index}페이지 OCR 실패: {exc}")
    parsed = parse_business_registration_text("\n".join(text_parts))
    if parsed.get("business_address"):
        messages.append("주소를 자동 추출했습니다. 정확한지 확인해 주세요.")
    if parsed.get("business_name"):
        messages.append("상호를 자동 추출했습니다.")
    if parsed.get("owner_name"):
        messages.append("대표자명을 자동 추출했습니다.")
    return parsed, messages or ["OCR을 실행했습니다. 결과를 확인해 주세요."], preview.getvalue()


def set_cell(sheet, coordinate: str, value: object) -> None:
    cell = sheet[coordinate]
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            cell = sheet.cell(merged_range.min_row, merged_range.min_col)
            break
    cell.value = value
    cell.alignment = copy(cell.alignment)
    cell.alignment = cell.alignment.copy(wrap_text=True, vertical="center")


def clear_yellow_fills(sheet) -> None:
    from openpyxl.styles import PatternFill

    no_fill = PatternFill(fill_type=None)
    for row in sheet.iter_rows():
        for cell in row:
            rgb = getattr(cell.fill.fgColor, "rgb", None)
            indexed = getattr(cell.fill.fgColor, "indexed", None)
            if rgb in YELLOW_FILL_RGB or indexed == 13:
                cell.fill = no_fill


def add_image_to_sheet(sheet, image_bytes: bytes | None, anchor: str, width: int, height: int) -> None:
    if not image_bytes:
        return
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from PIL import Image as PILImage

    image = PILImage.open(io.BytesIO(image_bytes))
    original_width, original_height = image.size
    scale = min(width / original_width, height / original_height) if original_width and original_height else 1
    output = io.BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    drawing = OpenpyxlImage(output)
    drawing.width = max(1, int(original_width * scale))
    drawing.height = max(1, int(original_height * scale))
    sheet.add_image(drawing, anchor)


def fill_contract_excel(inputs: ContractInputs) -> bytes:
    from openpyxl import load_workbook

    template_path = CONTRACT_TEMPLATES[inputs.contract_type]
    if template_path.exists():
        workbook = load_workbook(template_path)
    else:
        from contract_templates_embedded import get_template_bytes

        workbook = load_workbook(io.BytesIO(get_template_bytes(inputs.contract_type)))
    sheet = workbook["中文版"] if "中文版" in workbook.sheetnames else workbook.active
    suffix = re.sub(r"^NO\.?\s*KADOB", "", inputs.kadob_no.strip(), flags=re.IGNORECASE)
    duration_text = "自开通之日起【365】天" if inputs.contract_type == "1年" else "自开通之日起【3】个月"
    set_cell(sheet, "M1", f"NO.KADOB{suffix}" if suffix else "NO.KADOB")
    set_cell(sheet, "E4", inputs.store_name)
    set_cell(sheet, "J4", inputs.merchant_address)
    set_cell(sheet, "E5", inputs.owner_name)
    set_cell(sheet, "L5", "")
    set_cell(sheet, "C12", f"☐ 推广计划开始时间 {inputs.start_date.year}年 {inputs.start_date.month:02d}月{inputs.start_date.day:02d}日")
    set_cell(sheet, "D14", inputs.store_name)
    set_cell(sheet, "H14", inputs.city)
    set_cell(sheet, "B19", f"{inputs.store_name} / {inputs.shop_id}".strip(" /"))
    set_cell(sheet, "D19", inputs.city)
    set_cell(sheet, "L19", duration_text)
    if inputs.memo.strip():
        set_cell(sheet, "C54", f"{sheet['C54'].value or ''}\n补充备注：{inputs.memo.strip()}")
    add_image_to_sheet(sheet, inputs.signature_bytes, "D58", 110, 38)
    add_image_to_sheet(sheet, inputs.stamp_bytes, "F57", 72, 72)
    clear_yellow_fills(sheet)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def contract_pdf(inputs: ContractInputs) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    font_name = "Helvetica"
    for path in ["/System/Library/Fonts/AppleSDGothicNeo.ttc", "/System/Library/Fonts/Hiragino Sans GB.ttc"]:
        try:
            if Path(path).exists():
                pdfmetrics.registerFont(TTFont("ContractFont", path))
                font_name = "ContractFont"
                break
        except Exception:
            pass
    if font_name == "Helvetica":
        pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
        font_name = "HYGothic-Medium"

    styles = getSampleStyleSheet()
    title = ParagraphStyle("Title", parent=styles["Title"], fontName=font_name, fontSize=18, leading=24, alignment=TA_LEFT)
    rows = [
        ["계약기간", inputs.contract_type],
        ["NO.KADOB", inputs.kadob_no],
        ["商户名称 / 门店1", inputs.store_name],
        ["商户地址", inputs.merchant_address],
        ["经营者/法定代表人姓名", inputs.owner_name],
        ["所在城市", inputs.city],
        ["Shopid", inputs.shop_id or "-"],
        ["推广计划开始时间", inputs.start_date.isoformat()],
    ]
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=16 * mm, rightMargin=16 * mm, topMargin=16 * mm, bottomMargin=16 * mm)
    table = Table(rows, colWidths=[48 * mm, 128 * mm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D8D3C4")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    doc.build([Paragraph("계약서 생성본", title), Spacer(1, 5 * mm), table])
    output.seek(0)
    return output.getvalue()


def zip_outputs(inputs: ContractInputs, excel_bytes: bytes, pdf_bytes: bytes) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(f"{safe_filename(inputs.store_name)}_계약서.xlsx", excel_bytes)
        archive.writestr(f"{safe_filename(inputs.store_name)}_계약서.pdf", pdf_bytes)
    output.seek(0)
    return output.getvalue()


def show_dependency_status() -> None:
    missing = []
    for module, label in [("openpyxl", "Excel 생성"), ("reportlab", "PDF 생성"), ("PIL", "이미지 처리"), ("pytesseract", "OCR")]:
        try:
            __import__(module)
        except Exception:
            missing.append(label)
    if missing:
        st.warning("일부 선택 기능을 사용할 수 없습니다: " + ", ".join(missing) + ". 앱은 수동 입력 방식으로 계속 사용할 수 있습니다.")


def render_app(embedded: bool = False) -> None:
    if not embedded:
        st.set_page_config(page_title="계약서 자동 생성", page_icon="📄", layout="wide")
    st.markdown(
        """
        <style>
        div[data-baseweb="input"] input,
        div[data-baseweb="textarea"] textarea,
        div[data-baseweb="select"] > div {
            background-color: #fff8b8 !important;
        }
        div[role="radiogroup"] {
            background-color: #fff8b8;
            border-radius: 8px;
            padding: 8px 10px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    if not embedded:
        st.title("계약서 자동 생성")
        st.caption("1年/3个月 계약서 템플릿을 선택하고 입력값을 반영해 Excel/PDF를 생성합니다.")
    show_dependency_status()

    missing_templates = [label for label, path in CONTRACT_TEMPLATES.items() if not path.exists()]
    if missing_templates:
        try:
            from contract_templates_embedded import TEMPLATES as EMBEDDED_TEMPLATES

            missing_templates = [label for label in missing_templates if label not in EMBEDDED_TEMPLATES]
        except Exception:
            pass
    if missing_templates:
        st.error(f"계약서 템플릿 파일이 없습니다: {', '.join(missing_templates)}")
        return

    left, right = st.columns([1.1, 0.9], gap="large")
    with left:
        st.subheader("사업자등록증 및 기본 정보")
        business_license = st.file_uploader("사업자등록증 파일", type=["pdf", "png", "jpg", "jpeg", "heic"])
        if business_license:
            marker = f"{business_license.name}:{business_license.size}"
            if st.session_state.get("ocr_marker") != marker:
                parsed, messages, preview_bytes = ocr_business_registration(business_license.getvalue(), business_license.name)
                st.session_state["ocr_marker"] = marker
                st.session_state["ocr_messages"] = messages
                st.session_state["ocr_preview_bytes"] = preview_bytes
                st.session_state["ocr_business_name"] = parsed.get("business_name", "")
                st.session_state["ocr_owner_name"] = parsed.get("owner_name", "")
                st.session_state["ocr_business_address"] = parsed.get("business_address", "")
            for message in st.session_state.get("ocr_messages", []):
                st.info(message)
            with st.expander("OCR 결과 확인/수정", expanded=True):
                if st.session_state.get("ocr_preview_bytes"):
                    st.image(st.session_state["ocr_preview_bytes"], caption="OCR 전 자동 보정된 이미지")
                ocr_business = st.text_input("OCR 상호", key="ocr_business_name")
                ocr_owner = st.text_input("OCR 대표자명", key="ocr_owner_name")
                ocr_address = st.text_area("OCR 사업장 주소", key="ocr_business_address")
                if st.button("OCR 결과를 입력값에 반영"):
                    if ocr_business:
                        st.session_state["store_name"] = ocr_business
                    if ocr_owner:
                        st.session_state["owner_name"] = romanize_korean(ocr_owner)
                    if ocr_address:
                        st.session_state["address_kr"] = ocr_address
                        st.session_state["merchant_address"] = auto_english_address(ocr_address)
                    if hasattr(st, "rerun"):
                        st.rerun()
                    elif hasattr(st, "experimental_rerun"):
                        st.experimental_rerun()

        contract_type = st.radio("계약기간", ["1年", "3个月"], horizontal=True)
        kadob_no = st.text_input("NO.KADOB 뒤 입력값", placeholder="예: 9030256")
        store_name = st.text_input("商户名称 / 门店1", key="store_name")
        address_kr = st.text_area("한국어 주소", key="address_kr")
        if address_kr and not st.session_state.get("merchant_address"):
            st.session_state["merchant_address"] = auto_english_address(address_kr)
        merchant_address = st.text_area("商户地址 (영문 주소)", key="merchant_address")
        owner_name = st.text_input("经营者/法定代表人姓名 (여권 영문명)", key="owner_name")
        city = st.selectbox("所在城市", CITY_OPTIONS)
        shop_id = st.text_input("Shopid / 매장코드", placeholder="선택 입력")
        start_date = st.date_input("推广计划开始时间", value=date.today())

    with right:
        st.subheader("서명 및 출력")
        ad_product = "템플릿 원본 유지"
        st.info("계약서 템플릿에 있는 금액과 상품/套餐 영역은 원본 그대로 유지합니다.")
        memo = st.text_area("비고")
        signature_file = st.file_uploader("甲方 서명 이미지", type=["png", "jpg", "jpeg"])
        stamp_file = st.file_uploader("甲方 도장 이미지", type=["png", "jpg", "jpeg"])
        output_type = st.radio("출력 형식", ["Excel (.xlsx)", "PDF (.pdf)", "Excel + PDF"], horizontal=False)
        generate = st.button("계약서 생성", type="primary")

    if not generate:
        return

    missing = []
    if not kadob_no.strip():
        missing.append("NO.KADOB")
    if not store_name.strip():
        missing.append("商户名称")
    if not merchant_address.strip():
        missing.append("商户地址")
    if not owner_name.strip():
        missing.append("经营者/法定代表人姓名")
    if missing:
        st.error("필수 항목이 누락되었습니다: " + ", ".join(missing))
        return

    inputs = ContractInputs(
        contract_type=contract_type,
        kadob_no=kadob_no.strip(),
        store_name=store_name.strip(),
        merchant_address=merchant_address.strip(),
        owner_name=owner_name.strip(),
        city=city,
        shop_id=shop_id.strip(),
        start_date=start_date,
        ad_product=ad_product.strip(),
        memo=memo.strip(),
        signature_bytes=signature_file.getvalue() if signature_file else None,
        stamp_bytes=stamp_file.getvalue() if stamp_file else None,
    )

    try:
        with st.spinner("계약서를 생성하는 중입니다..."):
            excel_bytes = fill_contract_excel(inputs)
            pdf_bytes = contract_pdf(inputs)
    except Exception as exc:
        st.error("계약서 생성 중 오류가 발생했습니다.")
        st.exception(exc)
        return

    st.success("생성 완료")
    if output_type == "Excel (.xlsx)":
        st.download_button("Excel 다운로드", excel_bytes, file_name=f"{safe_filename(inputs.store_name)}_계약서.xlsx")
    elif output_type == "PDF (.pdf)":
        st.download_button("PDF 다운로드", pdf_bytes, file_name=f"{safe_filename(inputs.store_name)}_계약서.pdf")
    else:
        st.download_button("Excel + PDF 다운로드", zip_outputs(inputs, excel_bytes, pdf_bytes), file_name=f"{safe_filename(inputs.store_name)}_계약서.zip")


if __name__ == "__main__":
    render_app()
